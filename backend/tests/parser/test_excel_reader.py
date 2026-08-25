from pathlib import Path

import pytest
from app.parser.excel_reader import WorkbookLimitError, read_workbook
from openpyxl import Workbook


def _save_workbook(
    path: Path, *, sheets: int = 1, rows: int = 1, columns: int = 1
) -> None:
    workbook = Workbook()
    workbook.active.title = "Sheet 1"
    for sheet_index in range(1, sheets):
        workbook.create_sheet(f"Sheet {sheet_index + 1}")
    for sheet in workbook.worksheets:
        for row in range(1, rows + 1):
            for column in range(1, columns + 1):
                sheet.cell(row=row, column=column, value="value")
    workbook.save(path)


def test_xlsx_sheet_count_limit_is_enforced(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "too-many-sheets.xlsx"
    _save_workbook(path, sheets=2)
    monkeypatch.setattr("app.parser.excel_reader.MAX_WORKSHEETS", 1)

    with pytest.raises(WorkbookLimitError, match="workbook_too_many_worksheets"):
        read_workbook(path)


@pytest.mark.parametrize(
    ("rows", "columns", "limit_name", "error_code"),
    [
        (3, 1, "MAX_ROWS_PER_WORKSHEET", "workbook_too_many_rows"),
        (1, 3, "MAX_COLUMNS_PER_WORKSHEET", "workbook_too_many_columns"),
    ],
)
def test_xlsx_row_and_column_limits_are_enforced(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    rows: int,
    columns: int,
    limit_name: str,
    error_code: str,
) -> None:
    path = tmp_path / "oversized.xlsx"
    _save_workbook(path, rows=rows, columns=columns)
    monkeypatch.setattr(f"app.parser.excel_reader.{limit_name}", 2)

    with pytest.raises(WorkbookLimitError, match=error_code):
        read_workbook(path)
