"""Adapters for the two Excel formats accepted by the intake module."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class WorksheetData:
    name: str
    rows: tuple[tuple[object, ...], ...]


MAX_WORKSHEETS = 32
MAX_ROWS_PER_WORKSHEET = 100_000
MAX_COLUMNS_PER_WORKSHEET = 256


class WorkbookLimitError(ValueError):
    """Raised when a workbook exceeds the parser's resource limits."""


def read_workbook(path: Path) -> tuple[WorksheetData, ...]:
    """Read values only, keeping the parser independent from workbook objects."""

    extension = path.suffix.lower()
    if extension == ".xls":
        return _read_xls(path)
    if extension == ".xlsx":
        return _read_xlsx(path)
    raise ValueError("unsupported_extension")


def _read_xls(path: Path) -> tuple[WorksheetData, ...]:
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        if workbook.nsheets > MAX_WORKSHEETS:
            raise WorkbookLimitError("workbook_too_many_worksheets")
        worksheets: list[WorksheetData] = []
        for index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(index)
            if sheet.nrows > MAX_ROWS_PER_WORKSHEET:
                raise WorkbookLimitError("workbook_too_many_rows")
            if sheet.ncols > MAX_COLUMNS_PER_WORKSHEET:
                raise WorkbookLimitError("workbook_too_many_columns")
            worksheets.append(
                WorksheetData(
                    sheet.name,
                    tuple(
                        tuple(
                            sheet.cell_value(row, column)
                            for column in range(sheet.ncols)
                        )
                        for row in range(sheet.nrows)
                    ),
                )
            )
        return tuple(worksheets)
    finally:
        workbook.release_resources()


def _read_xlsx(path: Path) -> tuple[WorksheetData, ...]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.worksheets) > MAX_WORKSHEETS:
            raise WorkbookLimitError("workbook_too_many_worksheets")
        worksheets: list[WorksheetData] = []
        for sheet in workbook.worksheets:
            if (sheet.max_row or 0) > MAX_ROWS_PER_WORKSHEET:
                raise WorkbookLimitError("workbook_too_many_rows")
            if (sheet.max_column or 0) > MAX_COLUMNS_PER_WORKSHEET:
                raise WorkbookLimitError("workbook_too_many_columns")
            rows: list[tuple[object, ...]] = []
            for row_number, row in enumerate(
                sheet.iter_rows(values_only=True), start=1
            ):
                if row_number > MAX_ROWS_PER_WORKSHEET:
                    raise WorkbookLimitError("workbook_too_many_rows")
                if len(row) > MAX_COLUMNS_PER_WORKSHEET:
                    raise WorkbookLimitError("workbook_too_many_columns")
                rows.append(tuple(row))
            worksheets.append(WorksheetData(sheet.title, tuple(rows)))
        return tuple(worksheets)
    finally:
        workbook.close()
