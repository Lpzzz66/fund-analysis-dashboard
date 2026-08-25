"""Excel 元数据提取：估值表识别、估值日期、表头定位、候选产品与文本数字样本。

设计要点：
- ``xlrd`` 读取 ``.xls``、``openpyxl`` 只读流式读取 ``.xlsx``，两者都归一为
  ``SheetGrid``（工作表名 + 二维单元格数组），识别逻辑只依赖网格内容，
  不依赖固定行号，也不依赖具体 Excel 库。
- 识别逻辑（日期解析、表头定位、产品候选）是纯函数，可用内存网格直接单元测试。
- 单元格取值保持原始类型：文本、数字、``datetime/date``；文本归一化仅去首尾空白。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

# —— 估值表内容识别关键词（不依赖固定行号）——
HEADER_CODE = "科目代码"
HEADER_NAME = "科目名称"
DATE_LABEL = "估值日期"
# 净值类关键词：命中即增强“疑似估值表”判定
NAV_KEYWORDS: tuple[str, ...] = (
    "基金资产净值",
    "基金单位净值",
    "累计单位净值",
    "单位净值",
)
# 判定估值表时，除双表头外还需命中的强特征
NAV_STRONG_KEYWORDS: tuple[str, ...] = (
    "基金资产净值",
    "基金单位净值",
    "累计单位净值",
)

# 交易记录工作簿识别线索
TRANSACTION_HINTS: tuple[str, ...] = ("交易记录",)

# 默认产品别名（可扩展：ProductCatalog 接受任意别名序列，目录后缀规则独立于别名）
DEFAULT_PRODUCT_ALIASES: tuple[str, ...] = ("梦一号", "千金一号", "天策上将")
PRODUCT_DIR_SUFFIX = "估值表"

# 每张工作表参与关键词扫描的最大行数（.xlsx 流式读取的安全上限；
# 本目录估值表的表头在第 4 行、净值汇总区约在第 90-130 行，上限远大于此）
MAX_SCAN_ROWS_PER_SHEET = 500
# 表头行之前视为“标题区”的行数上限
MAX_TITLE_ROWS = 8
# 文本数字样本最多记录条数
MAX_TEXT_NUMBER_SAMPLES = 3

# —— 日期解析 ——
_DATE_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)"),  # 20260402
    re.compile(
        r"(?<!\d)(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})(?!\d)"
    ),  # 2026-04-02 / 2026/04/02
    re.compile(r"(?<!\d)(20\d{2})年(\d{1,2})月(\d{1,2})日?(?!\d)"),  # 2026年4月2日
)
# 明显的文本数字（千分位或带小数），例如 "10,824,713.18" / "3.74" / "-0.9719"
_TEXT_NUMBER_RE = re.compile(r"^-?\d{1,3}(,\d{3})+\.\d+$|^-?\d+\.\d+$")


def parse_date_text(text: str) -> date | None:
    """从任意文本中提取第一个合法日期；支持 20260402 / 2026-04-02 / 2026/04/02 / 2026年4月2日。"""
    if not text:
        return None
    for pattern in _DATE_PATTERNS:
        m = pattern.search(text)
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                continue
    return None


def cell_text(value: Any) -> str:
    """把单元格原始值归一为文本：None→空串，日期→ISO，其余→str 并去首尾空白。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # 科目代码等在 .xls 中常以浮点存储（1002.0），归一成 "1002" 便于识别
        return str(int(value))
    return str(value).strip()


@dataclass
class SheetGrid:
    """归一化后的工作表网格。rows[r][c] 为单元格原始值。"""

    name: str
    rows: list[list[Any]] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def col_count(self) -> int:
        return max((len(r) for r in self.rows), default=0)


@dataclass
class SheetFacts:
    """从一张工作表提取的结构化事实。"""

    sheet_name: str
    header_row: int | None = None  # 1-based；通过“科目代码+科目名称”定位
    has_code_header: bool = False
    has_name_header: bool = False
    has_date_label: bool = False
    valuation_date: date | None = None
    nav_hits: dict[str, int] = field(default_factory=dict)
    text_number_count: int = 0
    text_number_samples: list[str] = field(default_factory=list)
    title_lines: list[str] = field(default_factory=list)  # 表头行之前的非空文本行

    @property
    def is_valuation(self) -> bool:
        """估值表判定：双表头（科目代码+科目名称）为必要条件，
        另需命中估值日期或任一强净值关键词。"""
        if not (self.has_code_header and self.has_name_header):
            return False
        return self.has_date_label or any(
            k in self.nav_hits for k in NAV_STRONG_KEYWORDS
        )


@dataclass
class WorkbookFacts:
    """整个工作簿的合并事实：取第一张判定为估值表的工作表，否则取第一张。"""

    sheet_names: list[str] = field(default_factory=list)
    chosen_sheet: str | None = None
    is_valuation: bool = False
    facts: SheetFacts | None = None

    @property
    def valuation_date(self) -> date | None:
        return self.facts.valuation_date if self.facts else None

    def title_text(self) -> str:
        return " ".join(self.facts.title_lines) if self.facts else ""


class ProductCatalog:
    """候选产品目录：别名子串匹配 + “XX估值表”目录后缀规则，均可扩展。"""

    def __init__(self, aliases: tuple[str, ...] = DEFAULT_PRODUCT_ALIASES) -> None:
        self.aliases = tuple(aliases)

    def candidates_from_text(self, text: str) -> list[str]:
        """文本中命中的别名（按别名定义顺序、去重）。"""
        if not text:
            return []
        return [a for a in self.aliases if a in text]

    def candidates_from_dir_name(self, dir_name: str) -> list[str]:
        """目录名识别：先匹配别名子串，再把“<产品>估值表”后缀剥掉作为候选。"""
        hits = self.candidates_from_text(dir_name)
        if hits:
            return hits
        if dir_name.endswith(PRODUCT_DIR_SUFFIX) and len(dir_name) > len(
            PRODUCT_DIR_SUFFIX
        ):
            stripped = dir_name[: -len(PRODUCT_DIR_SUFFIX)].strip()
            if stripped:
                return [stripped]
        return []

    def resolve(self, values: list[str]) -> tuple[str | None, bool]:
        """去重后唯一 → (该值, False)；多个 → (None, True) 冲突；空 → (None, False)。"""
        unique = sorted(set(values))
        if not unique:
            return None, False
        if len(unique) == 1:
            return unique[0], False
        return None, True


def extract_facts(grid: SheetGrid) -> SheetFacts:
    """从网格提取事实：表头行、估值日期、净值关键词、文本数字、标题区。

    估值日期优先级：
    1. 含“估值日期”标签的单元格内文本中的日期；
    2. 同一行右侧相邻单元格（文本日期或日期类型单元格）；
    3. 前 10 行内第一个日期类型（datetime/date）单元格。
    """
    facts = SheetFacts(sheet_name=grid.name)
    scan_rows = grid.rows[:MAX_SCAN_ROWS_PER_SHEET]
    date_from_label: date | None = None
    date_from_typed_cell: date | None = None

    for r0, row in enumerate(scan_rows):
        row_texts = [cell_text(v) for v in row]
        # —— 表头行定位：同一行同时出现“科目代码”与“科目名称” ——
        if (
            facts.header_row is None
            and HEADER_CODE in row_texts
            and HEADER_NAME in row_texts
        ):
            facts.header_row = r0 + 1
            facts.has_code_header = True
            facts.has_name_header = True
        for c0, text in enumerate(row_texts):
            if not text:
                continue
            # —— 净值关键词计数 ——
            for kw in NAV_KEYWORDS:
                if kw in text:
                    facts.nav_hits[kw] = facts.nav_hits.get(kw, 0) + 1
            # —— 估值日期 ——
            if DATE_LABEL in text:
                facts.has_date_label = True
                if date_from_label is None:
                    d = parse_date_text(text.split(DATE_LABEL, 1)[1])
                    if d is None:
                        d = _date_from_right_neighbors(scan_rows, r0, c0 + 1)
                    if d is not None:
                        date_from_label = d
            # —— 文本数字（仅字符串型单元格）——
            if isinstance(row[c0], str) and _TEXT_NUMBER_RE.match(text):
                facts.text_number_count += 1
                if len(facts.text_number_samples) < MAX_TEXT_NUMBER_SAMPLES:
                    facts.text_number_samples.append(text)
        # —— 日期类型单元格兜底（前 10 行）——
        if date_from_typed_cell is None and r0 < 10:
            for v in row:
                if isinstance(v, datetime):
                    date_from_typed_cell = v.date()
                    break
                if isinstance(v, date):
                    date_from_typed_cell = v
                    break

    facts.valuation_date = date_from_label or date_from_typed_cell
    # —— 标题区：表头行之前的非空文本行（无表头则取前 MAX_TITLE_ROWS 行）——
    title_end = (
        (facts.header_row - 1)
        if facts.header_row
        else min(len(scan_rows), MAX_TITLE_ROWS)
    )
    for row in scan_rows[:title_end]:
        line = " ".join(t for t in (cell_text(v) for v in row) if t)
        if line:
            facts.title_lines.append(line)
    return facts


def _date_from_right_neighbors(
    scan_rows: list[list[Any]], r0: int, start_c0: int
) -> date | None:
    """估值日期标签右侧最多 3 个单元格内找文本日期或日期类型单元格。"""
    if r0 >= len(scan_rows):
        return None
    row = scan_rows[r0]
    for v in row[start_c0 : start_c0 + 3]:
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        d = parse_date_text(cell_text(v))
        if d is not None:
            return d
    return None


def analyze_grids(grids: list[SheetGrid]) -> WorkbookFacts:
    """合并多张工作表：优先取第一张判定为估值表的表。"""
    if not grids:
        return WorkbookFacts()
    all_facts = [extract_facts(g) for g in grids]
    chosen = next((f for f in all_facts if f.is_valuation), all_facts[0])
    return WorkbookFacts(
        sheet_names=[g.name for g in grids],
        chosen_sheet=chosen.sheet_name,
        is_valuation=chosen.is_valuation,
        facts=chosen,
    )


# ---------------------------------------------------------------------------
# 读取适配器：xlrd (.xls) / openpyxl (.xlsx)
# ---------------------------------------------------------------------------


def load_grids(path: Path) -> list[SheetGrid]:
    """按扩展名选择适配器读取工作簿；ImportError 原样抛出以便记为依赖缺失。"""
    ext = path.suffix.lower()
    if ext == ".xls":
        return _grids_from_xlrd(path)
    if ext in (".xlsx", ".xlsm"):
        return _grids_from_openpyxl(path)
    raise ValueError(f"unsupported excel extension: {ext!r}")


def _grids_from_xlrd(path: Path) -> list[SheetGrid]:
    import xlrd  # xlrd 2.x 仅支持 .xls

    wb = xlrd.open_workbook(str(path))
    grids: list[SheetGrid] = []
    for sh in wb.sheets():
        rows: list[list[Any]] = []
        for r in range(sh.nrows):
            row: list[Any] = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, wb.datemode)
                    except (ValueError, TypeError):
                        pass  # 非法日期序号保留原值
                row.append(value)
            rows.append(row)
        grids.append(SheetGrid(name=sh.name, rows=rows))
    return grids


def _grids_from_openpyxl(path: Path) -> list[SheetGrid]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    grids: list[SheetGrid] = []
    try:
        for ws in wb.worksheets:
            rows: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_SCAN_ROWS_PER_SHEET:
                    break
                rows.append(list(row))
            grids.append(SheetGrid(name=ws.title, rows=rows))
    finally:
        wb.close()
    return grids
