"""测试公共设施：合成估值表/交易记录 .xlsx、内存假 xlrd 对象、FileInfo 快速构造。

测试绝不依赖真实历史目录 ``F:\\AgentWorks\\估值表A``，也绝不修改任何源文件。
"""

from __future__ import annotations

import sys
from datetime import date as date_cls
from pathlib import Path

import pytest
from openpyxl import Workbook

# 让测试既能 `pytest tools/...` 直接运行，也能从任意 cwd 运行
_TOOLS_ROOT = Path(__file__).resolve().parents[3]
if str(_TOOLS_ROOT) not in sys.path:
    sys.path.insert(0, str(_TOOLS_ROOT))

from tools.valuation_inventory.models import (
    FileInfo,
    FileType,
    ParseStatus,
    SourceZone,
)


def make_valuation_xlsx(
    path: Path,
    *,
    title: str = "丹寅梦一号私募证券投资基金",
    date_text: str = "估值日期：20260402",
    header_row: int = 4,
    sheet_name: str | None = None,
    nav: bool = True,
    text_numbers: bool = True,
    extra_amount: float | None = None,
) -> None:
    """生成一张合成估值表 .xlsx，结构模仿真实估值表：标题区 + 表头 + 科目行 + 净值汇总。"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name or title
    ws.cell(row=1, column=1, value="证券投资基金估值表")
    ws.cell(row=2, column=1, value=f"上海丹寅投资管理中心___{title}___专用表")
    ws.cell(row=3, column=1, value=date_text)
    header = [
        "科目代码",
        "科目名称",
        "数量",
        "单位成本",
        "成本",
        "成本占净值%",
        "市价",
        "市值",
    ]
    for c, name in enumerate(header, start=1):
        ws.cell(row=header_row, column=c, value=name)
    data = [
        ("1002", "银行存款", 640251.73),
        ("100201", "活期存款", 640237.34),
    ]
    r = header_row + 1
    for code, name, amount in data:
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=5, value=amount)
        r += 1
    if extra_amount is not None:
        ws.cell(row=r, column=1, value="1102")
        ws.cell(row=r, column=2, value="清算款")
        ws.cell(row=r, column=5, value=extra_amount)
        r += 1
    if nav:
        ws.cell(row=r + 1, column=1, value="基金资产净值:")
        ws.cell(row=r + 1, column=5, value=131532087.37)
        ws.cell(row=r + 2, column=1, value="基金单位净值：")
        ws.cell(row=r + 2, column=2, value=2.5882)
        ws.cell(row=r + 3, column=1, value="累计单位净值:")
        ws.cell(row=r + 3, column=2, value=3.5079)
    if text_numbers:
        ws.cell(row=r + 5, column=2, value="10,824,713.18")
        ws.cell(row=r + 6, column=2, value="3.74")
        ws.cell(row=r + 7, column=2, value="-0.9719")
    wb.save(path)


def make_transaction_xlsx(path: Path, *, live: bool = True) -> None:
    """生成合成交易记录 .xlsx（单位净值 + 交易记录两页，无科目代码表头）。"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "单位净值"
    ws.append(["日期", "单位净值", "累计单位净值"])
    ws.append(["2024-07-02", 0.993, 0.993])
    ws.append(["2024-07-03", 1.014, 1.014])
    ws2 = wb.create_sheet("交易记录")
    ws2.append(["日期", "证券代码", "业务", "金额"])
    if live:
        ws2.append(["2024-07-02", "600000", "买入", 100000])
    wb.save(path)


class FakeCell:
    def __init__(self, value, ctype) -> None:
        self.value = value
        self.ctype = ctype


class FakeSheet:
    """鸭子类型的 xlrd sheet。rows[r][c] = (value, ctype)。"""

    def __init__(self, name: str, rows: list[list[tuple[object, int]]]) -> None:
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(r) for r in rows), default=0)

    def cell(self, rowx: int, colx: int) -> FakeCell:
        row = self._rows[rowx]
        if colx < len(row):
            value, ctype = row[colx]
        else:
            value, ctype = "", 0
        return FakeCell(value, ctype)


class FakeBook:
    def __init__(self, sheets: list[FakeSheet], datemode: int = 0) -> None:
        self._sheets = sheets
        self.datemode = datemode

    def sheets(self) -> list[FakeSheet]:
        return self._sheets


def make_file_info(
    rel_path: str,
    *,
    zone: SourceZone = SourceZone.PRIMARY,
    product: str | None = "梦一号",
    valuation_date: date_cls | None = None,
    sha256: str | None = "a" * 64,
    is_valuation: bool = True,
    identity_conflict: bool = False,
) -> FileInfo:
    """dedup/report 单元测试用的最小 FileInfo。"""
    name = rel_path.rsplit("/", 1)[-1]
    return FileInfo(
        rel_path=rel_path,
        file_name=name,
        ext=Path(name).suffix.lower(),
        size_bytes=100,
        mtime="2026-08-26T00:00:00+00:00",
        zone=zone,
        sha256=sha256,
        product=product,
        identity_conflict=identity_conflict,
        file_type=FileType.VALUATION_XLS if is_valuation else FileType.UNKNOWN,
        is_valuation=is_valuation,
        valuation_date=valuation_date,
        parse_status=ParseStatus.OK,
    )


@pytest.fixture()
def synth_tree(tmp_path: Path) -> Path:
    """标准合成目录：主目录 + gz，含重复、冲突、gz 独有、交易记录与旁车文件。"""
    return build_synth_tree(tmp_path)


def build_synth_tree(tmp_path: Path) -> Path:
    """synth_tree 的非 fixture 版本，供测试内直接调用。"""
    root = tmp_path / "估值表A"
    primary = root / "梦一号估值表" / "2026年01-12月" / "2026年01月"
    gz = root / "gz" / "梦一号估值表" / "2026年01-12月" / "2026年01月"
    tx_dir = root / "千金一号估值表" / "交易记录备份"
    primary.mkdir(parents=True)
    gz.mkdir(parents=True)
    tx_dir.mkdir(parents=True)

    make_valuation_xlsx(
        primary / "梦一号 01月05日.xlsx", date_text="估值日期：20260105"
    )
    make_valuation_xlsx(
        primary / "梦一号 01月06日.xlsx", date_text="估值日期：2026-01-06"
    )
    # gz 与主目录同日同内容（字节级拷贝）
    make_valuation_xlsx(gz / "梦一号 01月05日.xlsx", date_text="估值日期：20260105")
    # gz 同日不同内容（冲突）
    make_valuation_xlsx(
        gz / "梦一号 01月06日.xlsx",
        date_text="估值日期：2026/01/06",
        extra_amount=123.45,
    )
    # gz 独有日期
    make_valuation_xlsx(gz / "梦一号 01月07日.xlsx", date_text="估值日期：20260107")
    # 交易记录与旁车
    make_transaction_xlsx(root / "千金一号估值表" / "千金一号_交易记录.xlsx")
    make_transaction_xlsx(
        tx_dir / "_backup_before_千金一号_2026-01-05.xlsx", live=False
    )
    (root / "千金一号估值表" / "千金一号_交易记录.xlsx.bak_cum_nav").write_bytes(
        b"fake-sidecar"
    )
    # 损坏的 xls：不影响其他文件解析
    (gz / "梦一号 01月08日.xls").write_bytes(b"not-a-real-excel-file")
    return root
